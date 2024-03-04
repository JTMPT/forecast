import os
from shutil import copy
from openpyxl import load_workbook
import pandas as pd
from functions import up_load_df
from find_file import find_files_with_pattern
import sys

def export_status_exists(create_forecast_basic_location, forecast_version_basic_folder_location, forecast, software_data_folder_location,client_data_folder_location, file_date):
    sys.path.insert(0, r'{}'.format(create_forecast_basic_location))
    from run_create_forecast_basic import run_notebook

    folder_path=r'{}\For_approval\Reference_tabels'.format(client_data_folder_location)
    pattern='TAZ_V'
    matching_files=find_files_with_pattern(r'{}\shp'.format(folder_path), pattern)

    forecast_2020=None

    # אם יש שכבות חדשות
    if len(matching_files) > 0:
            suffix = '.shp'
            filtered_files_strings = []
            for string in matching_files:
                if string.endswith(suffix):
                    filtered_files_strings.append(string)

            filepath=filtered_files_strings[0]

            #load excel file
            workbook = load_workbook(filename=r"{}\inputs_outputs.xlsx".format(create_forecast_basic_location))

            #open workbook
            sheet = workbook.active

            #modify the desired cell
            sheet["B4"] = folder_path
            sheet["B5"] = True
            sheet["B6"] = filepath

            #save the file
            workbook.save(filename=r"{}\inputs_outputs.xlsx".format(create_forecast_basic_location))
            # להריץ את קוד בסיס עם השכבות
            execution_result = run_notebook(r'{}\run_basic_from_ad_hoc.ipynb'.format(create_forecast_basic_location))
            print("ad-hoc-Notebook execution result:", execution_result)

            if execution_result == True:
                src_path = r'{}\forecast_2020_{}.xlsx'.format(folder_path, file_date)
                destination_path = r'{}\forecast_2020_{}_with_taz_changes.xlsx'.format(folder_path, file_date)

                copy(src_path, destination_path)
                os.remove(src_path)

                src_full_path = r'{}\2020_jtmt_forcast_full_{}.xlsx'.format(folder_path, file_date)
                destination_full_path = r'{}\2020_jtmt_forcast_full_{}_with_taz_changes.xlsx'.format(folder_path, file_date)

                copy(src_full_path, destination_full_path)
                os.remove(src_full_path)

                forecast_2020=up_load_df(r'{}'.format(folder_path), r'2020_jtmt_forcast_full_{}_with_taz_changes'.format(file_date))
                #  forecast_2020=up_load_df(r'{}\background_files'.format(software_data_folder_location),'2020_jtmt_forcast_full_230720')

                col=[]

                col_20=['Taz_num','Taz_name',
                'main_secto',
                'aprt_20', 'pop_without_dorms_yeshiva',
                'student_toddlers',
                'student_gov',
                'cbs_muni_student_left_by_pre_of_demand_left',
                'uni_students', 'student_dorms',
                'emp_from_uni_student',
                'student_yeshiva',
                'emp_okev',
                'emp_not_okev','student']
                 
                forecast_2020=pd.merge(forecast[col].reset_index(),forecast_2020[col_20],how='left',on='Taz_num').fillna(0)

                save_excel_path=r'{}\For_approval\{}_forecast_2020_For_approval.xlsx'.format(client_data_folder_location,file_date)

                forecast_2020[col_20].to_excel(save_excel_path,index=False)
            return forecast_2020[col_20]
    # אם אין שכבות חדשות
    else:
        #load excel file
        workbook = load_workbook(filename=r"{}\inputs_outputs.xlsx".format(create_forecast_basic_location))

        #open workbook
        sheet = workbook.active

        #modify the desired cell
        sheet["B4"] = forecast_version_basic_folder_location
        sheet["B5"] = False
        sheet["B6"] = r'{}\background_files'.format(create_forecast_basic_location)

        #save the file
        workbook.save(filename=r"{}\inputs_outputs.xlsx".format(create_forecast_basic_location))

        forecast_2020=up_load_df(r'{}\background_files'.format(software_data_folder_location),'2020_jtmt_forcast_full_230720')

        col=[]

        col_20=['Taz_num','Taz_name',
        'main_secto',
        'aprt_20', 'pop_without_dorms_yeshiva',
        'student_toddlers',
        'student_gov',
        'cbs_muni_student_left_by_pre_of_demand_left',
        'uni_students', 'student_dorms',
        'emp_from_uni_student',
        'student_yeshiva',
        'emp_okev',
        'emp_not_okev','student']

        forecast_2020=pd.merge(forecast[col].reset_index(),forecast_2020[col_20],how='left',on='Taz_num').fillna(0)

        save_excel_path=r'{}\For_approval\{}_forecast_2020_For_approval.xlsx'.format(client_data_folder_location,file_date)

        forecast_2020[col_20].to_excel(save_excel_path,index=False)
    return forecast_2020[col_20]

